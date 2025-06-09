import os
import time
import spacy
from spacy.util import minibatch, compounding
from spacy.training import Example
import openpyxl
import warnings

warnings.filterwarnings("ignore", category=DeprecationWarning)
warnings.filterwarnings("ignore")

try:
    spacy.require_gpu()
    print("Using GPU for SpaCy.")
except Exception:
    print("GPU not available; falling back to CPU.")

nlp = spacy.load("en_core_web_lg")

if "textcat" not in nlp.pipe_names:
    nlp.add_pipe("textcat", last=True)
textcat = nlp.get_pipe("textcat")


def load_training_data(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    training_data = []

    intents = [cell.value for cell in sheet[1] if cell.value]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        for intent, phrase in zip(intents, row):
            if phrase:
                training_data.append((str(phrase), intent))

    return training_data, intents


def calculate_accuracy(model, data):

    correct = 0
    total = 0
    for text, true_label in data:
        doc = model.make_doc(text)
        scores = model(doc).cats
        predicted_label = max(scores, key=scores.get)
        if predicted_label == true_label:
            correct += 1
        total += 1
    return correct / total if total > 0 else 0.0



current_dir = os.path.dirname(os.path.abspath(__file__))
training_file = os.path.join(current_dir, "training_data.xlsx")

train_data, labels = load_training_data(training_file)

print("=== Corpus Summary ===")
print(f"Total training examples : {len(train_data)}")
label_counts = {lab: 0 for lab in labels}
for _, intent in train_data:
    label_counts[intent] += 1
for lab, cnt in label_counts.items():
    print(f"  • {lab}: {cnt} examples")
print("======================\n")

for label in labels:
    textcat.add_label(label)


log_wb = openpyxl.Workbook()
log_sheet = log_wb.active
log_sheet.title = "Training Log"

log_sheet.append([
    "Epoch",
    "Loss (textcat)",
    "Accuracy",
    "Corpus Count",
    "Batch Sizes (this epoch)",
    "Learning Rate",
    "Epoch Duration (s)"
])

other_pipes = [pipe for pipe in nlp.pipe_names if pipe != "textcat"]
print(f"Disabling pipelines for training: {other_pipes}")
nlp.disable_pipes(*other_pipes)

# Begin training:
optimizer = nlp.begin_training()
NUM_EPOCHS = 13  # Change 

accuracies = []
losses = []
durations = []

for epoch in range(1, NUM_EPOCHS + 1):
    print(f"\n— Starting Epoch {epoch}/{NUM_EPOCHS} —")
    start_time = time.time()

    epoch_losses = {}
    batch_size_list = []

    # Create minibatches that compound from 32 → 128:
    batches = minibatch(train_data, size=compounding(64.0, 256, 1.002))

    for batch in batches:
        batch_size = len(batch)
        batch_size_list.append(batch_size)

        examples = []
        for text, annotation in batch:
            doc = nlp.make_doc(text)
            cats = {lbl: (lbl == annotation) for lbl in labels}
            example = Example.from_dict(doc, {"cats": cats})
            examples.append(example)

        nlp.update(examples, sgd=optimizer, losses=epoch_losses)

    # End of epoch: calculate metrics
    loss_value = epoch_losses.get("textcat", 0.0)
    accuracy = calculate_accuracy(nlp, train_data)
    epoch_duration = time.time() - start_time

    accuracies.append(accuracy)
    losses.append(loss_value)
    durations.append(epoch_duration)

    # Attempt to fetch learning rate (version‐dependent):
    try:
        lr = getattr(optimizer, "learn_rate", getattr(optimizer, "alpha", None))
        if lr is None:
            lr = "N/A"
    except Exception:
        lr = "N/A"

    print(f"Epoch {epoch} ► Loss(textcat) = {loss_value:.4f}  |  Accuracy = {accuracy:.4f}")
    print(f"   • Corpus size       : {len(train_data)}")
    print(f"   • Batches           : {len(batch_size_list)} batches → sizes = {batch_size_list}")
    print(f"   • Learning rate     : {lr}")
    print(f"   • Epoch duration    : {epoch_duration:.2f} seconds\n")

    # Log everything into the Excel sheet:
    log_sheet.append([
        epoch,
        float(loss_value),
        float(accuracy),
        len(train_data),
        ", ".join(str(s) for s in batch_size_list),
        lr,
        round(epoch_duration, 2)
    ])

# ─────────────────────────────────────────────────────────────────────────────
# Save the model and the training log:
nlp.to_disk(current_dir)
print("Model saved to:", current_dir)

log_path = os.path.join(current_dir, "training_log.xlsx")
log_wb.save(log_path)
print("Training log saved to:", log_path)

# (Optional) Plot accuracy and loss over epochs:
import matplotlib.pyplot as plt

plt.figure(figsize=(12, 6))

plt.subplot(1, 2, 1)
plt.plot(range(1, NUM_EPOCHS + 1), accuracies, marker="o")
plt.xlabel("Epoch")
plt.ylabel("Accuracy")
plt.title("Training Accuracy")

plt.subplot(1, 2, 2)
plt.plot(range(1, NUM_EPOCHS + 1), losses, marker="o", color="red")
plt.xlabel("Epoch")
plt.ylabel("Loss")
plt.title("Training Loss")

plt.tight_layout()
plt.show()

print("Done!")
